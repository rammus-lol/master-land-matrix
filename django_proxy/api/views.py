import io
import requests
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
import geopandas as gpd
import pandas as pd
from shapely.geometry import shape
import traceback
from pathlib import Path
from django.conf import settings
from land_matrix_function import export
from api.custom_service.spatial_function import  geom_constructor
from api.custom_service.table_function import table_constructor
from api.custom_service.pdf_report import build_pdf_report
from rest_framework.decorators import api_view
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse
from drf_spectacular.types import OpenApiTypes
from .serializers import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

@extend_schema(exclude=True)
@api_view(['GET'])
@csrf_exempt
def generic_proxy(request, endpoint):
    """
    Generic proxy for the landmatrix.org API.
    Routes /api/<endpoint> requests to https://landmatrix.org/api/<endpoint>.
    """
    try:
        url = f'https://landmatrix.org/api/{endpoint}'
        params = request.GET.dict()
        response = requests.get(url, params=params, timeout=10)
        return JsonResponse(
            response.json(),
            status=response.status_code,
            safe=False
        )
    except requests.exceptions.RequestException as e:
        return JsonResponse(
            {'error': str(e)},
            status=500
        )
@csrf_exempt
@extend_schema(
    summary="Spatial analysis and geometry processing",
    description="""
        Processes a standard GeoJSON FeatureCollection  in EPSG:3857 along with precision flags.
        ...
        """,
    request=SpatialProcessSerializer,
    responses={
        200: GeomResponseSerializer,
    },
    tags=['geography']
)
@api_view(['POST'])
def geom(request):
    input_serializer = SpatialProcessSerializer(data=request.data)

    if not input_serializer.is_valid():
        return Response(input_serializer.errors, status=400)

    # Récupération des données validées
    geojson_data = input_serializer.validated_data['geojson']
    is_precise = input_serializer.validated_data['is_precise']
    try:
        query = gpd.GeoDataFrame.from_features(
            geojson_data["features"],
            crs="EPSG:3857"
        )
        test=query.get("radius")#->test if user asked for circular form
        if test is not None:
            gdf_circle=query[query.geometry.geom_type == 'Point']
            gdf_polygon=query[query.geometry.geom_type != 'Point']
            buffer=gdf_circle.buffer(gdf_circle['radius'])
            gdf_circle["geometry"]=buffer
            query=gpd.GeoDataFrame(pd.concat([gdf_circle,gdf_polygon],ignore_index=True),crs='EPSG:3857')
        spatial_query,number_of_deals=geom_constructor(query,is_precise)
        if type(spatial_query) is str:
            if spatial_query=="code_2":
                return JsonResponse({"status": "No deal inside the provided shapes but some are near by"
                                     ,"data": 0}, status=200)
            elif spatial_query=="code_1":
                return JsonResponse({"status": "No deal found",
                                     "data":0}, status=200)
        else:
            export = spatial_query.to_json()
            response = {"status": f"Spatial selection successful. Number of deals found: {number_of_deals}",
                        'data': json.loads(export)}
            return JsonResponse(response,status=200)

    except Exception as e:
        print("error in geom api querying :", e, flush=True)
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)
@csrf_exempt
@extend_schema(
    summary="Requesting data export (spreadsheet or PDF)",
    description="""
        Processes a list of IDs and exports the corresponding deals into either:
        - spreadsheet format (Excel or CSV)
        - PDF report with summary charts based on counts.
    """,
    request=SheetInputSerializer,
    responses={
        200: OpenApiResponse(
            description="The generated spreadsheet file",
            response=OpenApiTypes.BINARY,
        ),
        400: OpenApiResponse(description="Invalid IDs or format provided"),
    },
    tags=["spreadsheet file"]
)
@api_view(['POST'])
def sheet(request):
    payload = request.data if hasattr(request, "data") else request
    id_list = payload.get('id_list', [])
    file_format = payload.get('file_format') or payload.get('format')

    if not isinstance(id_list, list) or len(id_list) == 0:
        return JsonResponse({"error": "id_list must be a non-empty list"}, status=400)

    try:
        id_list = [int(deal_id) for deal_id in id_list]
    except (TypeError, ValueError):
        return JsonResponse({"error": "id_list must contain integers"}, status=400)

    if file_format not in ["xlsx", "csv", "pdf"]:
        return JsonResponse({"error": "format must be 'xlsx', 'csv' or 'pdf'"}, status=400)

    table = table_constructor(id_list)
    if file_format == "xlsx":
        output = io.BytesIO()
        table.to_excel(output, index=False, engine='openpyxl', sheet_name='Deals')
        
        # Adjust column widths based on content
        output.seek(0)
        workbook = load_workbook(output)
        worksheet = workbook.active
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        workbook.save(output)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "export.xlsx"
        data = output.getvalue()
    elif file_format == "csv":
            output = io.StringIO()
            table.to_csv(output, sep=';', index=False)
            content_type = "text/csv"
            filename = "export.csv"
            data = output.getvalue()
    else:
        data = build_pdf_report(table)
        content_type = "application/pdf"
        filename = "export_report.pdf"
    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

