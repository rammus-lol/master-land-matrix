import requests
import geopandas as gpd
from shapely import Point, Polygon
from openpyxl import load_workbook #library for excel files manipulation, I use it for styling cells according to the level of accuracy
from openpyxl.styles import PatternFill
import pandas as pd
import json
import os
def buffer_creation(geom:gpd.GeoDataFrame,is_agricultural=False):#is _agricultural is a future parameter for dyniamic buffer radius
    geom.to_crs(epsg=3857,inplace=True)
    zoning=gpd.GeoDataFrame(geometry=geom.buffer(10500),crs=3857) 
    return zoning 
"""I'm using a radius a little bit bigger than 10 km,
making sure it's englobing at least 10 km circle and not 9945 meters because of shape's weirdness."""
def eval_geom(buffer,points): #spatial querying 
    result=points.sjoin(buffer, predicate="within")
    result=result.drop(columns="index_right")
    return result if not result.empty else "The database stocks no points within the given shape"
def api_call(adress): #create the point geodataframe object
    data=requests.get(adress).json()
    extract=[]
    crs_validation=set()
    None_deals=[]
    for d in data:
        location=d["selected_version"]["locations"]
        for l in location:
            temp={}
            temp["nid"]=l.get("nid")
            if l.get("point") not in [None,{},"null"]:
                temp["crs"]=l.get("point").get("crs").get("properties").get("name")
                crs_validation.add(temp["crs"])
                temp["lon"]=l.get("point").get("coordinates")[0]
                temp["lat"]=l.get("point").get("coordinates")[1]
            else:
                None_deals.append(d)
                json_dump_path=os.path.join(os.path.dirname(__file__),"incomplete_deals.json")#create a json for unlocated points whereever you download this .py
                print(f"Incomplete positioning data for the deal {d["id"]} a json dump will be created for further analysis")
                continue
            temp["accuracy"]=l.get("level_of_accuracy")
            extract.append(temp)
    if len(crs_validation)!=1:
        for e in extract:
            df=pd.DataFrame({'lon':[e["lon"]],'lat':[e["lat"]],'crs':[e["crs"]]})#pandas asks for an iterable object as column
            print(df)
            if e["crs"]!="EPSG:4326":
                gdf=gpd.GeoDataFrame(df,geometry=gpd.points_from_xy(df["lon"],df["lat"],crs=df["crs"].iloc[0]))
                gdf.to_crs(4326)
                # x=e["lon"]
                # y=e["lat"]
                e["lon"]=gdf.geometry.x.iloc[0]
                e["lat"]=gdf.geometry.y.iloc[0]
                e["crs"]="EPSG:4326"
                # print(f"coordinates changed : {[x,y]}>>{[e["lon"],e["lat"]]}")
                df=pd.DataFrame(extract)
                gdf=gpd.GeoDataFrame(df,geometry=gpd.points_from_xy(df['lon'],df["lat"]),crs=df['crs'].iloc[0])
                gdf.to_crs(3857,inplace=True)
                gdf.drop(columns="crs",inplace=True)               
            else:
                continue
    else:
        df=pd.DataFrame(extract)
        gdf=gpd.GeoDataFrame(df,geometry=gpd.points_from_xy(df['lon'],df["lat"]),crs=df['crs'].iloc[0])
        gdf.to_crs(3857,inplace=True)
        gdf.drop(columns="crs",inplace=True)
    with open(json_dump_path,"w",encoding="utf-8") as f:
        None_deals.insert(0,{"url_used":adress})
        json.dump(None_deals,f,indent=4)
    return gdf
def export(gdataframe,destination:str="A:/cours_m2/land_matrix/found_land_matrix_deals.xlsx"):#Takes a geodataframe and export it to an excel file with colored cells according to the level of accuracy
    df=gdataframe.drop(columns="geometry")
    df.to_excel(destination,index=False)#export to excel the future document downloaded by users
    ref={"COUNTRY":PatternFill(start_color="8E0500", end_color="8E0500", fill_type="solid"),#dark red
"ADMINISTRATIVE_REGION":PatternFill(start_color="BF212F", end_color="BF212F", fill_type="solid"),#light red
"APPROXIMATE_LOCATION":PatternFill(start_color="F9A73E", end_color="F9A73E", fill_type="solid"),#dark yellow/light orange
"EXACT_LOCATION":PatternFill(start_color="006F3C", end_color="006F3C", fill_type="solid"),#dark green
"COORDINATES":PatternFill(start_color="3760E3", end_color="3760E3", fill_type="solid")}#medium blue
    #I classified as the api doc classifies level of accuracy
    wb=load_workbook(destination)
    ws=wb.active
    header=[cell.value for cell in ws[1]]
    accuracy_index=header.index("accuracy")+1#openpyxl uses 1 based indexing
    for row in range(2,ws.max_row+1):
        cell=ws.cell(row=row,column=accuracy_index)
        if cell.value in ref:
            cell.fill=ref[cell.value]
    i=2
    c_index=ws.max_column+1
    for k,v in ref.items():       
        cell=ws.cell(row=i,column=c_index)
        cell.value=k
        cell.fill=v
        i+=1
    wb.save(destination)              

